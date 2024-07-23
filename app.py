from flask import Flask, render_template, request, send_file, redirect, url_for, flash
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font

app = Flask(__name__)
app.secret_key = 'supersecretkey'

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/create_excel', methods=['POST'])
def create_excel():
    try:
        trip_group_name = request.form['trip_group_name']
        num_routes = int(request.form['num_routes'])
        
        routes = []
        route_numbers = []
        for i in range(num_routes):
            route_number = request.form[f'route_no_{i}']
            route_numbers.append(route_number)
            stop_names_text = request.form[f'stop_names_{i}']
            stop_names = stop_names_text.split('\n')
            start_time = request.form[f'start_time_{i}']
            end_time = request.form[f'end_time_{i}']
            ac_type = request.form[f'ac_type_{i}']
            
            num_time_zones = int(request.form[f'num_time_zones_{i}'])
            time_schedule_zones = []
            for tz in range(num_time_zones):
                num_trips = int(request.form[f'num_trips_{i}_{tz}'])
                interval = request.form[f'interval_{i}_{tz}']
                time_schedule_zones.append((num_trips, interval))
                
            routes.append(((stop_names, (start_time, end_time)), time_schedule_zones, ac_type))
        
        if 'stop_id_file' not in request.files:
            flash('No file part')
            return redirect(url_for('index'))
        
        stop_id_file = request.files['stop_id_file']
        
        if stop_id_file.filename == '':
            flash('No selected file')
            return redirect(url_for('index'))
        
        stop_id_df = pd.read_excel(stop_id_file)

        excel_data = create_bus_schedule(routes, stop_id_df, trip_group_name, route_numbers)

        return send_file(BytesIO(excel_data), download_name='Bus_Schedule_with_Trip_Group_and_Stops.xlsx', as_attachment=True)

    except Exception as e:
        print(f"Error: {str(e)}")  # Debug statement
        flash(str(e))
        return redirect(url_for('index'))

def create_bus_schedule(routes, stop_id_df, trip_group_name, route_numbers):
    columns = [
        "Trip_group", "route_id", "service_id", "direction_id", "trip_headsign",
        "stop_id", "stop_name", "stop_headsign", "pickup_type", "drop_off_type",
        "continuous_pickup", "continuous_drop_off", "timed_stop", "wait_time",
        "timepoint", "wheelchair_accessible", "bikes_allowed", "T01", "T02", "T03",
        "T04", "T05", "T06", "T07", "T08", "T09", "T10", "T11", "T12", "T13", "T14",
        "T15", "T16", "T17", "T18", "T19", "T20", "T21", "T22", "T23", "T24", "T25",
        "T26", "T27", "T28", "T29", "T30", "T31", "T32"
    ]

    route_columns = [
        "route_id", "agency_id", "route_short_name", "route_long_name", "route_desc",
        "route_type", "route_url", "route_color", "route_text_color", "route_sort_order",
        "continuous_pickup", "continuous_drop_off"
    ]

    df = pd.DataFrame(columns=columns)
    route_df = pd.DataFrame(columns=route_columns)

    for route_num, (route, time_schedule_zones, ac_type), route_number in zip(range(1, len(routes) + 1), routes, route_numbers):
        stop_names, start_end_times = route
        stop_names = [name.strip() for name in stop_names if name.strip()]
        num_stops = len(stop_names)

        start_time_str, end_time_str = start_end_times
        start_time = datetime.strptime(start_time_str, "%H:%M:%S")
        end_time = datetime.strptime(end_time_str, "%H:%M:%S")

        total_time = (end_time - start_time).total_seconds()
        interval_seconds = total_time / (num_stops - 1)
        interval = timedelta(seconds=interval_seconds)

        stops_with_times = [(stop_names[i], (start_time + i * interval).strftime("%H:%M:%S")) for i in range(num_stops)]

        if stops_with_times:
            first_stop_initials = ''.join([word[0] for word in stops_with_times[0][0].split()[:2]]).upper()
            last_stop_initials = ''.join([word[0] for word in stops_with_times[-1][0].split()[:2]]).upper()
            if ac_type == "AC":
                route_id = (first_stop_initials[0] + last_stop_initials[0] + "AC")
                route_long_name = f"{stops_with_times[0][0]} - {stops_with_times[-1][0]} AC"
            else:
                route_id = (first_stop_initials + last_stop_initials)[:4]
                route_long_name = f"{stops_with_times[0][0]} - {stops_with_times[-1][0]}"
        else:
            route_id = "UNKNOWN"
            route_long_name = "UNKNOWN"

        df.loc[len(df)] = {"Trip_group": f"MBMC{route_number}"}

        service_id = "FULLW"
        timed_stop = 1
        timepoint = 1

        for i, (stop_name, stop_time) in enumerate(stops_with_times):
            stop_id = stop_id_df.loc[stop_id_df['stop_name'] == stop_name, 'stop_id'].values
            if stop_id.size > 0:
                stop_id_value = stop_id[0]
            else:
                stop_id_value = None

            entry = {
                'route_id': route_id,
                'service_id': service_id,
                'stop_id': stop_id_value,
                'stop_name': stop_name,
                'timed_stop': timed_stop,
                'timepoint': timepoint,
                'T01': stop_time,
            }

            current_time = datetime.strptime(stop_time, "%H:%M:%S")

            trip_index = 1

            for tz_index, (num_trips, interval) in enumerate(time_schedule_zones):
                interval_delta = datetime.strptime(interval, "%H:%M:%S") - datetime(1900, 1, 1)

                if tz_index == 0 and num_trips == 1:
                    trips_to_calculate = num_trips

                elif tz_index == 0:
                    trips_to_calculate = num_trips - 1

                else:
                    trips_to_calculate = num_trips

                for _ in range(trips_to_calculate):
                    next_time = (current_time + interval_delta).strftime("%H:%M:%S")

                    trip_index += 1

                    entry[f'T{trip_index:02d}'] = next_time

                    current_time += interval_delta

            df = pd.concat([df, pd.DataFrame([entry])], ignore_index=True)

        route_entry = {
            "route_id": route_id,
            "agency_id": "apnawahan-mumbai-in",
            "route_short_name": route_number,
            "route_long_name": route_long_name,
            "route_desc": "",
            "route_type": 3,
            "route_url": "",
            "route_color": "",
            "route_text_color": "",
            "route_sort_order": "",
            "continuous_pickup": "",
            "continuous_drop_off": ""
        }
        route_df = pd.concat([route_df, pd.DataFrame([route_entry])], ignore_index=True)

        note_row = pd.Series(
            {"Trip_group": "Note: a row must be skipped between trip_groups. No data can entered in this row."})
        df = pd.concat([df, pd.DataFrame([note_row])], ignore_index=True)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Schedule', index=False)
        route_df.to_excel(writer, sheet_name='Routes', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Schedule']

        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for i, row in df.iterrows():
            if "Note:" in str(row["Trip_group"]):
                for col in range(1, len(columns) + 1):
                    cell = worksheet.cell(row=i + 2, column=col)
                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                    cell.font = Font(color="FFFFFF")
            else:
                for col in range(18, 50):
                    cell = worksheet.cell(row=i + 2, column=col)
                    cell.fill = green_fill

    return output.getvalue()

if __name__ == '__main__':
    app.run(debug=True)
